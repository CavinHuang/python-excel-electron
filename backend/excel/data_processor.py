from openpyxl import load_workbook, Workbook
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

from copy import copy
import traceback
# import logging
from .fetch_main_Image import fetch_img, destory_browser
from sse import logger

template_config = {
  "SEA_RAIL": {
    "merage_column": 13,
    "weight_column": 11,
    "pic_column": 14,
    "url_column": 17,
  },
  "AIR_TRANSPORT": {
    "merage_column": 15,
    "weight_column": 13,
    "pic_column": 21,
    "url_column": 20,
  }
}

# 设置单元格的对齐方式，使内容溢出时隐藏
alignment = Alignment(wrapText=True, shrinkToFit=False, wrap_text=True, vertical='center', horizontal='center')

def xlsx_sheet_copy(src_ws, targ_ws, logger):
    max_row = src_ws.max_row  # 最大行数
    max_column = src_ws.max_column  # 最大列数
    w, h = 0, 0
    #复制每个单元格
    logger.info('复制数据样式start')
    for column in range(1, max_column + 1):
        for row in range(1, max_row + 1):
            column_n = get_column_letter(column)
            i = '%s%d' % (column_n, row)  # 单元格编号
            try:
                #复制
                targ_ws[i].value = copy(src_ws[i].value)
                targ_ws[i].font = Font(src_ws[i].font.name, src_ws[i].font.size) #copy(src_ws[i].font)
                targ_ws[i].border = copy(src_ws[i].border)
                targ_ws[i].fill = copy(src_ws[i].fill)
                targ_ws[i].number_format = copy(src_ws[i].number_format)
                targ_ws[i].protection = copy(src_ws[i].protection)
                targ_ws[i].alignment = alignment #copy(src_ws[i].alignment)
            except Exception as e :
                print(e)

    logger.info('复制数据样式end')

    logger.info('处理合并单元格start')
    wm = list(src_ws.merged_cells)  # 开始处理合并单元格
    for i in range(0, len(wm)):
        cell2 = str(wm[i]).replace('(<MergedCellRange ', '').replace('>,)', '')
        targ_ws.merge_cells(cell2)

    logger.info('处理合并单元格end')

    logger.info('处理合并单元格行高列宽start')
    #此处有坑当你获得一个列宽为13的时候实际上是这个列和前面单元格一样的宽度，并不是他真的是13
    for i in range(1, max_column + 1):
        column_letter = get_column_letter(i)
        rs = src_ws.column_dimensions[column_letter].width
        if (rs == 13):
            rs = w
        else:
            w = rs
        targ_ws.column_dimensions[column_letter].width = rs
    #复制行高，没有列宽的坑
    for i in range(1, max_row + 1):
        targ_ws.row_dimensions[i].height = 20
    logger.info('处理合并单元格行高列宽end')
    logger.info('复制图片start')
    # 复制图片
    for img in src_ws._images:
      targ_ws.add_image(img)
    logger.info('复制图片start')

def xlsx_sheet_add_img(src_img, m, n, ws, logger):
    # 检查图片文件是否存在
    if os.path.exists(src_img):
        # 创建图片对象
        img = Image(src_img)
        # 设置图片的起始位置
        _from = AnchorMarker(n-1, 50000, m-1, 50000)
        # 设置图片的结束位置
        to = AnchorMarker(n, -50000, m, -50000)
        # 将图片锚定到指定位置
        img.anchor = TwoCellAnchor('twoCell', _from, to)
        # 将图片添加到工作表中
        ws.add_image(img)
        # 记录日志，表示图片插入成功
        logger.info('图片插入成功')

def process_excel(file_path, isPrice, isFetchImg, fileType):
    config = template_config[fileType]
    try:
        # logger = logging.getLogger('log')
        # logger.setLevel(logging.DEBUG)
        logger.info("开始处理Excel文件：%s", file_path)

        workbook = load_workbook(file_path)
        sheet = workbook.active

        new_workbook = Workbook()

        new_sheet = new_workbook.create_sheet(title="sheet 1")

        # 复制旧表格内容到第一个sheet
        # xlsx_sheet_copy(sheet, copy_sheet)

        # 获取合并单元格的范围
        merged_ranges = sheet.merged_cells.ranges

        sheet_images = sheet._images

        def has_image(rowIndex):
          for image in sheet_images:
            image_anchor = image.anchor
            if image.anchor._from.row == rowIndex:
               return True
          return False


        logger.info('==============复制表格数据 Start==============')

        # 在表格的最后新增一列，用于存放价格
        sheet.insert_cols(sheet.max_column + 1)
        sheet.cell(row=1, column=sheet.max_column, value='价格')

        # 遍历每一行数据
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            in_merged_range = False  # 标记当前单元格是否在合并单元格范围内
            merged_range = None  # 初始化合并单元格范围

            logger.info(f'开始处理第{row_idx}行数据：{row}')

            # print(row_idx)
            image_in_cell = has_image(row_idx)

            cell = sheet.cell(row_idx, config["url_column"])
            product_url = cell.value
            logger.info(f'商品链接:{product_url}')
            print(product_url)
            if product_url is not None and product_url.lower().startswith('http'):
                logger.info(f'=======没有图片: {row_idx}')
                file_full_name, price = fetch_img(url=product_url, row_index=row_idx, logger=logger, image_in_cell=image_in_cell, isPrice=isPrice, isFetchImg=isFetchImg)

                if image_in_cell is not True and isFetchImg is True:
                  xlsx_sheet_add_img(file_full_name, cell.row, config["pic_column"], new_sheet, logger)
                if isPrice is True:
                  new_sheet.cell(row=row_idx, column=sheet.max_column, value=price)
            else:
                logger.info(f'=======没有商品链接或者商品链接错误: {row_idx}')

            # 设置行高为20
            new_sheet.row_dimensions[row_idx].height = 20
            for col_idx, cell in enumerate(row, start=1):
                current_cell = sheet.cell(row=row_idx, column=col_idx)
                font = Font(name=current_cell.font.name, size=current_cell.font.size, vertAlign=current_cell.font.vertAlign, bold=current_cell.font.bold, italic=current_cell.font.italic, underline=current_cell.font.underline, strike=current_cell.font.strike, outline=current_cell.font.outline, shadow=current_cell.font.shadow, condense=current_cell.font.condense, extend=current_cell.font.extend, u=current_cell.font.u, charset=current_cell.font.charset, family=current_cell.font.family, scheme=current_cell.font.scheme)
                # 复制列宽
                if row_idx == 1:
                    col_letter = get_column_letter(col_idx)
                    new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[col_letter].width
                if col_idx == 1:
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    if value is None:
                        break

                # 检查当前单元格是否在合并单元格范围内
                for merged_range in merged_ranges:
                    if merged_range.min_row <= row_idx <= merged_range.max_row and \
                      merged_range.min_col <= col_idx <= merged_range.max_col:
                        in_merged_range = True
                        break
                if in_merged_range:
                    logger.info(f'第{row_idx}行数据需要拆分.')
                    # print(f'第{row_idx}行数据需要拆分.')
                    # 如果在合并单元格范围内，获取合并单元格的左上角单元格的值
                    # if col_idx == 13 :
                    if col_idx == config["merage_column"]:
                      if row_idx >= merged_range.max_row:
                          # 统计所有的数量
                          totalNum = 0
                          merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value or 0

                          for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                              value = sheet.cell(row=r_idx, column=col_idx - 1).value or 0
                              totalNum = totalNum + float(value)

                          # 计算平均值
                          quantity = float(merged_value) / totalNum
                          logger.info(f'开始拆分[重量]，共合并{merged_range.max_row - merged_range.min_row}行，总数量：{totalNum},总重量：{merged_value},平均值：{quantity}')


                          for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                              # 当前行的数量
                              num = sheet.cell(row=r_idx, column=col_idx - 1).value or 0
                              # 当前行的总重量
                              totalWeight = round(quantity * num, 2)
                              logger.info(f'当前行:{r_idx},数量：{num},当前项的平均值：{quantity}, 当前行总重量：{totalWeight}')
                              _cell = new_sheet.cell(row=r_idx, column=col_idx, value=totalWeight)
                              _cell.alignment = alignment
                              _cell.font = font
                    # 箱子
                    # elif col_idx == 11:
                    elif col_idx == config["weight_column"]:
                      if row_idx >= merged_range.max_row:
                        for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                          _cell = new_sheet.cell(row=r_idx, column=col_idx, value=0)
                          _cell.alignment = alignment
                          _cell.font = font

                        merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value or 0
                        new_sheet.cell(row=merged_range.min_row, column=merged_range.min_col, value=merged_value)

                    else:
                        # new_row.append(cell)
                        merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value or 0
                        _cell = new_sheet.cell(row=row_idx, column=col_idx, value=merged_value)
                        _cell.alignment = alignment
                        _cell.font = font

                    in_merged_range = False  # 重置标记
                else:
                    # 如果不在合并单元格范围内，直接使用当前单元格的值
                    # new_row.append(cell)
                    _cell = new_sheet.cell(row=row_idx, column=col_idx, value=cell)
                    _cell.alignment = alignment
                    _cell.font = font
                    in_merged_range = False  # 重置标记

        # 复制图片
        logger.info('==================复制表格数据 End==================')
        logger.info('==================复制表格图片 Start==================')
        for image in sheet._images:
          new_sheet.add_image(image)

        logger.info('==================复制表格图片 End==================')

        file_name, file_extension = os.path.splitext(os.path.basename(file_path))
        new_file_name = f"{file_name}_拆分表{file_extension}"
        new_file_path = os.path.dirname(file_path) +'/'+ new_file_name

        new_workbook.save(new_file_path)

        # 关闭工作簿
        workbook.close()
        new_workbook.close()
        destory_browser()

        logger.info('==================复制汇总表 Start==================')
        copy_total_sheet(file_path, new_file_path, logger)
    except Exception as e:
        print("处理出错：")
        print(traceback.format_exc())

def copy_total_sheet(file_path, src_path, logger):
    try:
      workbook = load_workbook(file_path)
      src_book = load_workbook(src_path)

      sheet = workbook.active
      src_sheet = src_book.active

      # 复制旧表格内容到第一个sheet
      xlsx_sheet_copy(sheet, src_sheet, logger)

      logger.info('===============复制汇总表 End================')

      file_name, file_extension = os.path.splitext(os.path.basename(file_path))
      new_file_name = f"{file_name}_拆分表{file_extension}"
      desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
      new_file_path = desktop_path +'/'+ new_file_name


      logger.info(f'原文件地址: {file_path}')
      logger.info(f'新文件地址: {new_file_path}')

      src_book.save(new_file_path)

      # 关闭工作簿
      workbook.close()
      src_book.close()

      logger.info("complate: 拆分完成")
    except Exception as e:
      print("处理出错：")
      print(traceback.format_exc())